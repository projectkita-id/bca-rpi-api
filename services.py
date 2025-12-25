import os
import json
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from datetime import datetime
from fastapi import HTTPException
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ========================================
# STYLING CONSTANTS
# ========================================

HEADER_FILL = PatternFill(
    start_color="1F4E78",
    end_color="1F4E78",
    fill_type="solid"
)

SUBHEADER_FILL = PatternFill(
    start_color="D9E1F2",
    end_color="D9E1F2",
    fill_type="solid"
)

PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FALLBACK_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SUBHEADER_FONT = Font(bold=True, size=10, color="1F4E78")
NORMAL_FONT = Font(size=10)
PASS_FONT = Font(bold=True, color="006100")
FAIL_FONT = Font(bold=True, color="9C0006")

THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")


# ========================================
# DATA NORMALIZATION
# ========================================

def normalize_item(item: dict, record_id: int, scanner_used: list[int]) -> dict:
    """
    Normalize item data from database format.
    
    Args:
        item: Raw item from database
        record_id: Batch record ID
        scanner_used: List of active scanner IDs [1, 2, 3]
    
    Returns:
        Normalized item dict with all scanner data
    """
    def get(scanner_key):
        """Safely extract scanner value and valid status"""
        data = item.get(scanner_key)
        if not data:
            return None, None
        return data.get("value"), data.get("valid")
    
    # Extract all scanner data
    scanners = {
        1: get("scanner_1"),
        2: get("scanner_2"),
        3: get("scanner_3")
    }
    
    # Determine result
    result = "PASS"
    fallback = False
    
    for scanner in scanner_used:
        value, valid = scanners.get(scanner, (None, None))
        
        if value is None:
            fallback = True
        
        if valid is False:
            result = "FAIL"
            break
    
    return {
        "item_id": item["item_id"],
        "record_id": record_id,
        "scanner_1": scanners[1][0],
        "scanner_1_valid": scanners[1][1],
        "scanner_2": scanners[2][0],
        "scanner_2_valid": scanners[2][1],
        "scanner_3": scanners[3][0],
        "scanner_3_valid": scanners[3][1],
        "result": result,
        "fallback": fallback
    }


# ========================================
# EXCEL IMPORT (Excel → JSON Database)
# ========================================

def excel_to_json(file):
    """
    Convert Excel file to JSON database format.
    
    Args:
        file: Uploaded Excel file object
    
    Returns:
        Dict with status, items count, and database path
    """
    try:
        file.seek(0)
        df = pd.read_excel(file, dtype=str)
        
        # Normalize column headers (case-insensitive)
        normalized_cols = {
            c.lower().strip(): c for c in df.columns
        }
        
        # Validate required columns
        required = ["scanner 1", "scanner 2", "scanner 3"]
        for col in required:
            if col not in normalized_cols:
                raise HTTPException(
                    400,
                    f"Missing required column: '{col.upper()}'"
                )
        
        def clean(value):
            """Clean cell value"""
            if pd.isna(value):
                return None
            return str(value).strip()
        
        # Build result
        result = []
        for _, row in df.iterrows():
            result.append({
                "Scanner 1": clean(row[normalized_cols["scanner 1"]]),
                "Scanner 2": clean(row[normalized_cols["scanner 2"]]),
                "Scanner 3": clean(row[normalized_cols["scanner 3"]]),
            })
        
        # Save to database file
        db_path = os.path.expanduser("~/scanner-db.json")
        
        with open(db_path, "w", encoding="utf-8") as f:
            json.dump(result, f, indent=2, ensure_ascii=False)
        
        return {
            "status": "ok",
            "items": len(result),
            "path": db_path
        }
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Excel parsing error: {str(e)}")


# ========================================
# EXCEL EXPORT (JSON → Beautiful Excel)
# ========================================

def export_record_to_excel(record_id: int, items: list, scanner_used: list[int]):
    """
    Export batch data to professional Excel format.
    
    Args:
        record_id: Batch ID
        items: List of item dicts from database
        scanner_used: List of active scanner IDs [1, 2, 3]
    
    Returns:
        Tuple of (BytesIO file object, filename)
    """
    
    print(f"📊 Exporting Batch #{record_id}: {len(items)} items, Scanners: {scanner_used}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Batch {record_id}"
    
    # ========================================
    # SUMMARY SECTION (Top of Sheet)
    # ========================================
    
    # Calculate stats
    total_items = len(items)
    pass_count = sum(1 for item in items if item.get("result") == "PASS")
    fail_count = sum(1 for item in items if item.get("result") == "FAIL")
    fallback_count = sum(1 for item in items if item.get("fallback") == True)
    
    # Get timestamps
    start_time = items[0].get("created_at", "") if items else ""
    end_time = items[-1].get("created_at", "") if items else ""
    
    # Parse timestamps for duration
    duration = "N/A"
    if start_time and end_time:
        try:
            start_dt = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
            end_dt = datetime.fromisoformat(end_time.replace('Z', '+00:00'))
            delta = end_dt - start_dt
            minutes = int(delta.total_seconds() // 60)
            seconds = int(delta.total_seconds() % 60)
            duration = f"{minutes}m {seconds}s"
        except:
            pass
    
    # Title row
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"BATCH SCAN REPORT #{record_id}"
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 25
    
    # Summary section headers
    ws["A3"] = "Total Items"
    ws["B3"] = total_items
    ws["A4"] = "PASS"
    ws["B4"] = pass_count
    ws["A5"] = "FAIL"
    ws["B5"] = fail_count
    ws["A6"] = "Fallback"
    ws["B6"] = fallback_count
    
    ws["D3"] = "Start Time"
    ws["E3"] = start_time
    ws["D4"] = "End Time"
    ws["E4"] = end_time
    ws["D5"] = "Duration"
    ws["E5"] = duration
    ws["D6"] = "Scanners Used"
    ws["E6"] = ", ".join([f"S{s}" for s in scanner_used])
    
    # Style summary section
    for row in range(3, 7):
        ws[f"A{row}"].font = SUBHEADER_FONT
        ws[f"A{row}"].alignment = LEFT_ALIGN
        ws[f"D{row}"].font = SUBHEADER_FONT
        ws[f"D{row}"].alignment = LEFT_ALIGN
        
        # Color coding
        if ws[f"A{row}"].value == "PASS":
            ws[f"B{row}"].fill = PASS_FILL
            ws[f"B{row}"].font = PASS_FONT
        elif ws[f"A{row}"].value == "FAIL":
            ws[f"B{row}"].fill = FAIL_FILL
            ws[f"B{row}"].font = FAIL_FONT
        elif ws[f"A{row}"].value == "Fallback":
            ws[f"B{row}"].fill = FALLBACK_FILL
    
    # ========================================
    # DATA TABLE SECTION
    # ========================================
    
    # Start data table at row 9
    header_row = 9
    data_start_row = header_row + 1
    
    # Build dynamic headers based on scanner_used
    headers = ["No", "Item ID"]
    
    for scanner_num in scanner_used:
        headers.append(f"Scanner {scanner_num}")
        headers.append("Valid")
    
    headers.extend(["Result", "Timestamp"])
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # Set column widths
    ws.column_dimensions["A"].width = 6   # No
    ws.column_dimensions["B"].width = 12  # Item ID
    
    col_letter = 2
    for _ in scanner_used:
        col_letter += 1
        ws.column_dimensions[get_column_letter(col_letter)].width = 30  # Scanner value
        col_letter += 1
        ws.column_dimensions[get_column_letter(col_letter)].width = 8   # Valid
    
    col_letter += 1
    ws.column_dimensions[get_column_letter(col_letter)].width = 10  # Result
    col_letter += 1
    ws.column_dimensions[get_column_letter(col_letter)].width = 20  # Timestamp
    
    # ========================================
    # DATA ROWS
    # ========================================
    
    for row_idx, item in enumerate(items, start=data_start_row):
        col = 1
        
        # No
        cell = ws.cell(row=row_idx, column=col, value=row_idx - data_start_row + 1)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        col += 1
        
        # Item ID
        cell = ws.cell(row=row_idx, column=col, value=item.get("item_id", ""))
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        col += 1
        
        # Scanner data (dynamic based on scanner_used)
        for scanner_num in scanner_used:
            scanner_key = f"scanner_{scanner_num}"
            valid_key = f"scanner_{scanner_num}_valid"
            
            # Scanner value
            value = item.get(scanner_key, "")
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = LEFT_ALIGN
            cell.border = THIN_BORDER
            
            # Color based on validity
            valid = item.get(valid_key)
            if valid is True:
                cell.fill = PASS_FILL
            elif valid is False:
                cell.fill = FAIL_FILL
            elif value is None or value == "":
                cell.fill = FALLBACK_FILL
            
            col += 1
            
            # Valid status
            valid_symbol = "✓" if valid is True else "✗" if valid is False else "-"
            cell = ws.cell(row=row_idx, column=col, value=valid_symbol)
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            
            if valid is True:
                cell.font = PASS_FONT
            elif valid is False:
                cell.font = FAIL_FONT
            
            col += 1
        
        # Result
        result = item.get("result", "")
        cell = ws.cell(row=row_idx, column=col, value=result)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        cell.font = Font(bold=True)
        
        if result == "PASS":
            cell.fill = PASS_FILL
            cell.font = PASS_FONT
        else:
            cell.fill = FAIL_FILL
            cell.font = FAIL_FONT
        
        col += 1
        
        # Timestamp
        timestamp = item.get("created_at", "")
        if timestamp:
            try:
                dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                timestamp = dt.strftime("%Y-%m-%d %H:%M:%S")
            except:
                pass
        
        cell = ws.cell(row=row_idx, column=col, value=timestamp)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # ========================================
    # FREEZE PANES & AUTO-FILTER
    # ========================================
    
    ws.freeze_panes = f"C{data_start_row}"  # Freeze headers and first 2 columns
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{data_start_row + len(items) - 1}"
    
    # ========================================
    # SAVE TO MEMORY
    # ========================================
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = f"Batch_{record_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    print(f"✅ Excel exported: {filename}")
    
    return output, filename


# ========================================
# MODULE LOADED CONFIRMATION
# ========================================

print("✅ services.py loaded successfully")
